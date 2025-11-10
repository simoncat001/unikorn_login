#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Populate the synchrotron radiation template from real-world data drops.

The data delivered to :mod:`remote_trans` now consists of raw XAS/XES scan
files, logbook excerpts and assorted metadata sheets.  Only some experiments
ship a ready-to-consume ``metadata.json``/``.yaml`` file, while others rely on
plain-text records such as ``scan_info.txt`` or notebook exports.  This helper
ingests as much information as it can find in a dataset directory—including
the CIF headers embedded at the top of ``.cbf`` diffraction frames—maps the
values onto the ``同步辐射表征元数据规范-2025.json`` template and prints a filled
JSON structure.  When structured metadata files are present they take
precedence; otherwise the script falls back to heuristics that recognise
common key/value phrases (both English and Chinese) and timestamps extracted
from logbooks.

Usage
-----

.. code-block:: bash

    python synchrotron_metadata.py \
        remote_trans/data/synchrotron_radiation/<dataset_dir> \
        --pretty

The command accepts either the dataset directory or any file located inside the
drop (for example ``metadata.json``).  When a file path is supplied the helper
scans its parent directory to collect the remaining artefacts so that the
template's “原始文件” section can be populated automatically.
"""

from __future__ import annotations

import argparse
import datetime as _dt
import json
import re
import sys
import unicodedata
from copy import deepcopy
from pathlib import Path
from typing import Callable, Dict, Iterable, Iterator, List, MutableMapping, Tuple

try:  # pragma: no cover - optional dependency
    import yaml  # type: ignore
except Exception:  # pragma: no cover - PyYAML is optional
    yaml = None  # type: ignore

try:  # pragma: no cover - optional dependency
    from openpyxl import load_workbook  # type: ignore
except Exception:  # pragma: no cover - optional dependency
    load_workbook = None  # type: ignore

try:  # pragma: no cover - optional dependency
    import xlrd  # type: ignore
except Exception:  # pragma: no cover - optional dependency
    xlrd = None  # type: ignore

HERE = Path(__file__).resolve().parent
DEFAULT_TEMPLATE = HERE.parent / "templates" / "synchrotron_radiation" / "同步辐射表征元数据规范-2025.json"
STRUCTURED_SUFFIXES = {".json", ".yaml", ".yml"}
CBF_SUFFIXES = {".cbf"}
TEXTUAL_SUFFIXES = {
    ".txt",
    ".dat",
    ".cfg",
    ".ini",
    ".info",
    ".log",
    ".lst",
    ".csv",
    ".par",
    ".fio",
    ".xdi",
}
TABULAR_SUFFIXES_OPENPYXL = {".xlsx", ".xlsm"}
TABULAR_SUFFIXES_XLRD = {".xls"}
TABULAR_SUFFIXES = TABULAR_SUFFIXES_OPENPYXL | TABULAR_SUFFIXES_XLRD

RAW_ALIASES: Dict[str, Tuple[str, str]] = {
    "experiment name": ("experiment", "name"),
    "experiment title": ("experiment", "name"),
    "experiment": ("experiment", "name"),
    "实验名称": ("experiment", "name"),
    "principal investigator": ("experiment", "principal_investigator"),
    "pi": ("experiment", "principal_investigator"),
    "实验负责人": ("experiment", "principal_investigator"),
    "负责人": ("experiment", "principal_investigator"),
    "experiment team": ("experiment", "team"),
    "team": ("experiment", "team"),
    "实验团队": ("experiment", "team"),
    "location": ("experiment", "location"),
    "beamline location": ("experiment", "location"),
    "实验地点": ("experiment", "location"),
    "start time": ("experiment", "start_time"),
    "开始时间": ("experiment", "start_time"),
    "实验时间": ("experiment", "start_time"),
    "实验开始时间": ("experiment", "start_time"),
    "end time": ("experiment", "end_time"),
    "结束时间": ("experiment", "end_time"),
    "实验结束时间": ("experiment", "end_time"),
    "notes": ("experiment", "notes"),
    "实验备注": ("experiment", "notes"),
    "实验日期": ("experiment", "start_time"),
    "MGID": ("experiment", "mgid_custom"),
    "MGID 自定义部分": ("experiment", "mgid_custom"),
    "MGID自定义部分": ("experiment", "mgid_custom"),
    "MGID custom": ("experiment", "mgid_custom"),
    "MGID custom part": ("experiment", "mgid_custom"),
    "beamline": ("beamline", "beamline"),
    "beamline name": ("beamline", "beamline"),
    "束线名称": ("beamline", "beamline"),
    "束线": ("beamline", "beamline"),
    "束线号": ("beamline", "beamline"),
    "束线编号": ("beamline", "beamline"),
    "_synchrotron_beamline": ("beamline", "beamline"),
    "_synchrotron_beamline.name": ("beamline", "beamline"),
    "_synchrotron_source.beamline": ("beamline", "beamline"),
    "_diffrn_source.beamline": ("beamline", "beamline"),
    "facility": ("beamline", "facility"),
    "facility name": ("beamline", "facility"),
    "光源设施": ("beamline", "facility"),
    "光源设施名称": ("beamline", "facility"),
    "_synchrotron_source": ("beamline", "facility"),
    "_synchrotron_source.facility": ("beamline", "facility"),
    "_synchrotron_source.description": ("beamline", "facility"),
    "_diffrn_source.source": ("beamline", "facility"),
    "_diffrn_source.target": ("beamline", "facility"),
    "photon energy": ("beamline", "photon_energy_eV"),
    "光子能量": ("beamline", "photon_energy_eV"),
    "光子能量(ev)": ("beamline", "photon_energy_eV"),
    "光子能量(eV)": ("beamline", "photon_energy_eV"),
    "光子能量 (ev)": ("beamline", "photon_energy_eV"),
    "光子能量(e v)": ("beamline", "photon_energy_eV"),
    "photon energy (ev)": ("beamline", "photon_energy_eV"),
    "beam energy": ("beamline", "photon_energy_eV"),
    "wavelength": ("beamline", "photon_energy_eV"),
    "wavelength (a)": ("beamline", "photon_energy_eV"),
    "wavelength (angstrom)": ("beamline", "photon_energy_eV"),
    "incident energy": ("beamline", "photon_energy_eV"),
    "_diffrn_source.energy": ("beamline", "photon_energy_eV"),
    "_diffrn_radiation_wavelength": ("beamline", "photon_energy_eV"),
    "_diffrn_radiation_wavelength.wavelength": ("beamline", "photon_energy_eV"),
    "storage ring current": ("beamline", "storage_ring_current_mA"),
    "储存环电流": ("beamline", "storage_ring_current_mA"),
    "储存环电流(mA)": ("beamline", "storage_ring_current_mA"),
    "ring current": ("beamline", "storage_ring_current_mA"),
    "_diffrn_source.current": ("beamline", "storage_ring_current_mA"),
    "monochromator": ("beamline", "monochromator"),
    "mono name": ("beamline", "monochromator"),
    "单色器": ("beamline", "monochromator"),
    "单色器类型": ("beamline", "monochromator"),
    "exit slit": ("beamline", "exit_slit_um"),
    "exit slit width": ("beamline", "exit_slit_um"),
    "出射狭缝": ("beamline", "exit_slit_um"),
    "出射狭缝宽度": ("beamline", "exit_slit_um"),
    "出射狭缝宽度(mm)": ("beamline", "exit_slit_um"),
    "detector name": ("detector", "name"),
    "detector": ("detector", "name"),
    "探测器名称": ("detector", "name"),
    "_diffrn_detector.detector": ("detector", "name"),
    "detector model": ("detector", "model"),
    "探测器型号": ("detector", "model"),
    "detector type": ("detector", "model"),
    "_diffrn_detector.type": ("detector", "model"),
    "_diffrn_detector.model": ("detector", "model"),
    "detector distance": ("detector", "distance_mm"),
    "detector_distance": ("detector", "distance_mm"),
    "detector-distance": ("detector", "distance_mm"),
    "detector distance (mm)": ("detector", "distance_mm"),
    "sample to detector distance": ("detector", "distance_mm"),
    "样品到探测器距离": ("detector", "distance_mm"),
    "样品到探测器距离(mm)": ("detector", "distance_mm"),
    "_diffrn_detector.reference_distance": ("detector", "distance_mm"),
    "_diffrn_detector.distance": ("detector", "distance_mm"),
    "sample name": ("sample", "name"),
    "样品名称": ("sample", "name"),
    "sample": ("sample", "name"),
    "样品代号": ("sample", "name"),
    "样品编号": ("sample", "name"),
    "_sample.name": ("sample", "name"),
    "_sample.id": ("sample", "name"),
    "sample environment": ("sample", "environment"),
    "样品环境": ("sample", "environment"),
    "样品描述": ("sample", "environment"),
    "样品状态": ("sample", "environment"),
    "environment": ("sample", "environment"),
    "_sample.description": ("sample", "environment"),
    "_sample.details": ("sample", "environment"),
    "temperature": ("sample", "temperature_K"),
    "样品温度": ("sample", "temperature_K"),
    "temperature (k)": ("sample", "temperature_K"),
    "temperature (c)": ("sample", "temperature_K"),
    "temperature (°c)": ("sample", "temperature_K"),
    "样品温度(℃)": ("sample", "temperature_K"),
    "temperature /k": ("sample", "temperature_K"),
    "temperature / °c": ("sample", "temperature_K"),
    "环境温度": ("sample", "temperature_K"),
    "_diffrn_measurement.sample_temperature": ("sample", "temperature_K"),
    "_sample.temperature": ("sample", "temperature_K"),
    "scan mode": ("scan", "mode"),
    "扫描模式": ("scan", "mode"),
    "mode": ("scan", "mode"),
    "扫描方式": ("scan", "mode"),
    "扫描类型": ("scan", "mode"),
    "scan points": ("scan", "points"),
    "points": ("scan", "points"),
    "扫描点数": ("scan", "points"),
    "数据点数": ("scan", "points"),
    "能量点数": ("scan", "points"),
    "测点数": ("scan", "points"),
    "扫描步数": ("scan", "points"),
    "_scan.points": ("scan", "points"),
    "_diffrn_scan.points": ("scan", "points"),
    "dwell time": ("scan", "dwell_time_s"),
    "integration time": ("scan", "dwell_time_s"),
    "exposure time": ("scan", "dwell_time_s"),
    "count time": ("scan", "dwell_time_s"),
    "collection time": ("scan", "dwell_time_s"),
    "单点积分时间": ("scan", "dwell_time_s"),
    "单点积分时间(s)": ("scan", "dwell_time_s"),
    "积分时间": ("scan", "dwell_time_s"),
    "积分时间(s)": ("scan", "dwell_time_s"),
    "曝光时间": ("scan", "dwell_time_s"),
    "测量时间": ("scan", "dwell_time_s"),
    "_scan.integration_time": ("scan", "dwell_time_s"),
    "_diffrn_scan.integration_time": ("scan", "dwell_time_s"),
    "energy range": ("scan", "energy_range_eV"),
    "energy range (ev)": ("scan", "energy_range_eV"),
    "扫描能区范围": ("scan", "energy_range_eV"),
    "扫描能量范围": ("scan", "energy_range_eV"),
    "起始能量": ("scan", "_energy_start_eV"),
    "开始能量": ("scan", "_energy_start_eV"),
    "start energy": ("scan", "_energy_start_eV"),
    "start energy (ev)": ("scan", "_energy_start_eV"),
    "初始能量": ("scan", "_energy_start_eV"),
    "终止能量": ("scan", "_energy_end_eV"),
    "结束能量": ("scan", "_energy_end_eV"),
    "end energy": ("scan", "_energy_end_eV"),
    "end energy (ev)": ("scan", "_energy_end_eV"),
    "final energy": ("scan", "_energy_end_eV"),
    "scan axis": ("scan", "mode"),
    "scan_axis": ("scan", "mode"),
    "scan-axis": ("scan", "mode"),
    "comments": ("scan", "comments"),
    "备注": ("scan", "comments"),
    "关联样品MGID": ("sample", "mgid_list"),
    "关联样品 mgid": ("sample", "mgid_list"),
    "样品MGID": ("sample", "mgid_list"),
    "样品 mgid": ("sample", "mgid_list"),
    "原始数据文件": ("files", "primary"),
    "表征原始数据文件": ("files", "primary"),
    "primary data file": ("files", "primary"),
    "原始数据列表": ("files", "listing"),
    "表征原始数据列表": ("files", "listing"),
    "表征原始数据列表(相对路径)": ("files", "listing"),
    "data file list": ("files", "listing"),
    "timestamp": ("experiment", "start_time"),
    "start time stamp": ("experiment", "start_time"),
    "start timestamp": ("experiment", "start_time"),
    "recorded": ("experiment", "start_time"),
    "title": ("experiment", "name"),
}

NUMERIC_FIELDS = {
    ("beamline", "photon_energy_eV"),
    ("beamline", "storage_ring_current_mA"),
    ("beamline", "exit_slit_um"),
    ("detector", "distance_mm"),
    ("sample", "temperature_K"),
    ("scan", "points"),
    ("scan", "dwell_time_s"),
}

DATETIME_FIELDS = {
    ("experiment", "start_time"),
    ("experiment", "end_time"),
}

DATA_PRIORITIES = (".cbf", ".dat", ".h5", ".hdf5", ".nxs", ".nx", ".csv", ".txt")

TEMPLATE_FIELD_ALIASES = {
    "实验名称": ("experiment", "name"),
    "实验负责人": ("experiment", "principal_investigator"),
    "实验团队": ("experiment", "team"),
    "实验地点": ("experiment", "location"),
    "实验备注": ("experiment", "notes"),
    "实验日期": ("experiment", "start_time"),
    "MGID自定义部分": ("experiment", "mgid_custom"),
    "MGID": ("experiment", "mgid_custom"),
    "备注": ("scan", "comments"),
    "关联样品MGID": ("sample", "mgid_list"),
}

TEMPLATE_SECTION_FIELD_ALIASES = {
    ("光束参数", "光源设施名称"): ("beamline", "facility"),
    ("光束参数", "束线名称"): ("beamline", "beamline"),
    ("光束参数", "光子能量(eV)"): ("beamline", "photon_energy_eV"),
    ("光束参数", "光子能量(ev)"): ("beamline", "photon_energy_eV"),
    ("光束参数", "储存环电流(mA)"): ("beamline", "storage_ring_current_mA"),
    ("光束参数", "单色器"): ("beamline", "monochromator"),
    ("光束参数", "出射狭缝宽度(μm)"): ("beamline", "exit_slit_um"),
    ("探测系统", "探测器名称"): ("detector", "name"),
    ("探测系统", "探测器型号"): ("detector", "model"),
    ("探测系统", "样品到探测器距离(mm)"): ("detector", "distance_mm"),
    ("样品信息", "样品名称"): ("sample", "name"),
    ("样品信息", "样品环境"): ("sample", "environment"),
    ("样品信息", "样品温度(K)"): ("sample", "temperature_K"),
    ("扫描参数", "扫描模式"): ("scan", "mode"),
    ("扫描参数", "扫描点数"): ("scan", "points"),
    ("扫描参数", "单点积分时间(s)"): ("scan", "dwell_time_s"),
    ("扫描参数", "扫描能区范围(eV)"): ("scan", "energy_range_eV"),
    ("扫描参数", "扫描能量范围(eV)"): ("scan", "energy_range_eV"),
    ("原始文件", "表征原始数据文件"): ("files", "primary"),
    ("原始文件", "原始数据文件"): ("files", "primary"),
    ("原始文件", "表征原始数据列表"): ("files", "listing"),
    ("原始文件", "原始数据列表"): ("files", "listing"),
}


def _parse_datetime(value: str) -> _dt.datetime | None:
    value = value.strip()
    if not value:
        return None
    try:
        return _dt.datetime.fromisoformat(value)
    except ValueError:
        pass
    try:
        date_only = _dt.date.fromisoformat(value)
    except ValueError:
        pass
    else:
        return _dt.datetime.combine(date_only, _dt.time())
    for fmt in (
        "%Y-%m-%d %H:%M:%S",
        "%Y/%m/%d %H:%M:%S",
        "%Y-%m-%dT%H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y/%m/%d %H:%M",
        "%Y年%m月%d日 %H:%M:%S",
        "%Y年%m月%d日 %H:%M",
        "%Y年%m月%d日",
    ):
        try:
            return _dt.datetime.strptime(value, fmt)
        except ValueError:
            continue
    return None


def _update(target: Dict[str, object], key: str, value: object) -> None:
    if value not in (None, ""):
        target[key] = value


def _collect_dataset_files(input_path: Path) -> Tuple[Path, List[Path]]:
    if input_path.is_file():
        dataset_dir = input_path.parent
    elif input_path.is_dir():
        dataset_dir = input_path
    else:
        raise FileNotFoundError(f"{input_path} does not exist")
    files = sorted(p for p in dataset_dir.rglob("*") if p.is_file())
    if not files:
        raise FileNotFoundError(f"No files found in dataset directory {dataset_dir}")
    return dataset_dir, files


def _read_structured_metadata(path: Path) -> Dict[str, object]:
    text = path.read_text(encoding="utf-8")
    if path.suffix.lower() == ".json":
        data = json.loads(text)
        if not isinstance(data, dict):
            raise TypeError(f"Top-level structure in {path} must be a JSON object")
        return data
    if path.suffix.lower() in {".yaml", ".yml"}:
        if yaml is None:  # pragma: no cover - optional dependency path
            raise RuntimeError(
                f"Cannot parse {path.name}: install PyYAML to read YAML files"
            )
        data = yaml.safe_load(text)
        if not isinstance(data, dict):
            raise TypeError(f"Unexpected metadata structure in {path}")
        return data
    raise ValueError(f"Unsupported structured metadata format: {path.suffix}")


def _normalise_key(key: str) -> str:
    key = key.strip().strip(":=：")
    key = key.replace("（", "(").replace("）", ")")
    key = unicodedata.normalize("NFKC", key)
    key = re.sub(r"[./\\、]+", " ", key)
    key = re.sub(r"[\s_]+", " ", key)
    return key.strip().lower()


ALIASES: Dict[str, Tuple[str, str]] = {
    _normalise_key(key): value for key, value in RAW_ALIASES.items()
}


def _iter_text_key_values(path: Path) -> Iterator[Tuple[str, str]]:
    for raw_line in path.read_text(encoding="utf-8", errors="ignore").splitlines():
        line = raw_line.strip()
        if not line:
            continue
        line = line.lstrip("#;*").strip()
        if not line:
            continue
        key: str | None = None
        value: str | None = None
        if "\t" in line:
            key, value = line.split("\t", 1)
        elif ":" in line:
            key, value = line.split(":", 1)
        elif "=" in line:
            key, value = line.split("=", 1)
        else:
            comma_match = re.match(r"^([^,]+),(.+)$", line)
            if comma_match:
                candidate_key = comma_match.group(1).strip()
                candidate_value = comma_match.group(2).strip()
                if _normalise_key(candidate_key) in ALIASES:
                    key, value = candidate_key, candidate_value
            if key is None and re.search(r"\s{2,}", line):
                parts = re.split(r"\s{2,}", line, maxsplit=1)
                if len(parts) == 2 and re.search(r"[A-Za-z\u4e00-\u9fff]", parts[0]):
                    key, value = parts
            if key is None:
                tokens = line.split()
                if len(tokens) >= 2:
                    for split_index in range(len(tokens) - 1, 0, -1):
                        candidate_key = " ".join(tokens[:split_index])
                        candidate_value = " ".join(tokens[split_index:])
                        if not candidate_key or not candidate_value:
                            continue
                        alias_key = _normalise_key(candidate_key)
                        if alias_key in ALIASES:
                            key = candidate_key
                            value = candidate_value
                            break
        if key is None or value is None:
            continue
        key = key.strip()
        value = value.strip()
        if not key or not value:
            continue
        yield key, value


def _parse_numeric(value: str) -> float | int | None:
    cleaned = value.replace(",", " ")
    match = re.search(r"[-+]?\d+(?:\.\d+)?", cleaned)
    if not match:
        return None
    number = float(match.group(0))
    if number.is_integer():
        return int(number)
    return number


def _strip_cif_value(value: str) -> str:
    value = value.strip()
    if not value:
        return value
    if value.startswith("#"):
        return ""
    if "#" in value:
        value = value.split("#", 1)[0].strip()
    if value.startswith(";"):
        return value.lstrip(";").strip()
    if value[0] in {'"', "'"}:
        quote = value[0]
        if value.endswith(quote) and len(value) > 1:
            value = value[1:-1]
        else:
            value = value[1:]
    return value.strip()


def _read_cbf_text_header(path: Path, *, chunk_size: int = 65_536) -> str:
    marker = b"--CIF-BINARY-FORMAT-SECTION--"
    buffer = bytearray()
    with path.open("rb") as handle:
        while True:
            chunk = handle.read(chunk_size)
            if not chunk:
                break
            buffer.extend(chunk)
            if marker in buffer or b"\f" in chunk:
                break
    text = buffer.decode("latin-1", errors="ignore")
    for sentinel in ("--CIF-BINARY-FORMAT-SECTION--", "\f", "\x1a"):
        if sentinel in text:
            text = text.split(sentinel, 1)[0]
    return text


def _split_cbf_comment(comment: str) -> Tuple[str | None, str | None]:
    comment = comment.strip()
    if not comment:
        return None, None
    for separator in (":", "=", "\t"):
        if separator not in comment:
            continue
        if separator == ":" and ": " not in comment and " :" not in comment:
            continue
        key, value = comment.split(separator, 1)
        return key.strip(), value.strip()
    parts = re.split(r"\s{2,}", comment, maxsplit=1)
    if len(parts) == 2:
        return parts[0].strip(), parts[1].strip()
    tokens = comment.split(None, 1)
    if len(tokens) == 2:
        return tokens[0].strip(), tokens[1].strip()
    return comment.strip(), ""


def _parse_cbf_header(path: Path) -> Dict[str, str]:
    header = _read_cbf_text_header(path)
    values: Dict[str, str] = {}
    current_key: str | None = None
    buffer: List[str] = []
    collecting = False
    for raw_line in header.splitlines():
        line = raw_line.strip()
        if not line:
            continue
        if line.startswith("#"):
            comment_key, comment_value = _split_cbf_comment(raw_line.lstrip("#"))
            if comment_key:
                cleaned = _strip_cif_value(comment_value)
                if cleaned:
                    values[comment_key] = cleaned
            continue
        if line.lower().startswith("loop_"):
            if collecting and current_key:
                text_value = "\n".join(buffer).strip()
                if text_value and text_value not in {"?", "."}:
                    values[current_key] = text_value
            current_key = None
            buffer = []
            collecting = False
            continue
        if line.startswith("data_"):
            continue
        if line.startswith("_"):
            if collecting and current_key:
                text_value = "\n".join(buffer).strip()
                if text_value and text_value not in {"?", "."}:
                    values[current_key] = text_value
            parts = line.split(None, 1)
            current_key = parts[0]
            buffer = []
            if len(parts) == 1:
                collecting = True
                continue
            value_part = _strip_cif_value(parts[1])
            collecting = False
            if value_part and value_part not in {"?", "."}:
                values[current_key] = value_part
            current_key = None
            continue
        if collecting and current_key:
            if line == ";":
                text_value = "\n".join(buffer).strip()
                if text_value and text_value not in {"?", "."}:
                    values[current_key] = text_value
                current_key = None
                buffer = []
                collecting = False
            else:
                buffer.append(raw_line.rstrip())
    if collecting and current_key:
        text_value = "\n".join(buffer).strip()
        if text_value and text_value not in {"?", "."}:
            values[current_key] = text_value
    return values


def _cbf_energy_to_electron_volts(raw_value: str) -> str | None:
    numeric = _parse_numeric(raw_value)
    if numeric is None:
        return None
    value = float(numeric)
    upper = raw_value.upper()
    if "KEV" in upper or ("EV" not in upper and value < 100):
        value *= 1_000.0
    if value <= 0:
        return None
    if abs(value - round(value)) < 1e-6:
        value = float(round(value))
    return f"{int(value)} eV" if value.is_integer() else f"{value:.3f} eV"


def _cbf_wavelength_to_energy(raw_value: str) -> str | None:
    numeric = _parse_numeric(raw_value)
    if numeric is None:
        return None
    wavelength = float(numeric)
    if wavelength <= 0:
        return None
    # Assume Angstrom input (common for CBF headers)
    energy = 12_398.419843320025 / wavelength
    if abs(energy - round(energy)) < 1e-6:
        energy = float(round(energy))
    return f"{int(energy)} eV" if energy.is_integer() else f"{energy:.3f} eV"


CBF_TRANSFORMS: Dict[str, Callable[[str], str | None]] = {
    "_diffrn_source.energy": _cbf_energy_to_electron_volts,
    "_diffrn_source.wavelength": _cbf_wavelength_to_energy,
    "_diffrn_radiation_wavelength": _cbf_wavelength_to_energy,
    "_diffrn_radiation_wavelength.wavelength": _cbf_wavelength_to_energy,
    "_synchrotron_photon_wavelength": _cbf_wavelength_to_energy,
    "wavelength": _cbf_wavelength_to_energy,
    "wavelength (a)": _cbf_wavelength_to_energy,
    "wavelength (angstrom)": _cbf_wavelength_to_energy,
}


def _coerce_value(section: str, key: str, value: str, *, original_key: str | None = None) -> object:
    if (section, key) in DATETIME_FIELDS:
        parsed = _parse_datetime(value)
        return parsed.isoformat() if parsed else value
    if (section, key) in NUMERIC_FIELDS or (section, key) == ("sample", "temperature_K"):
        numeric = _parse_numeric(value)
        if numeric is None:
            return value
        upper = value.upper()
        if (section, key) == ("beamline", "photon_energy_eV"):
            if "KEV" in upper:
                return numeric * 1_000
        if (section, key) == ("sample", "temperature_K"):
            if "℃" in value or "摄氏" in value or "°C" in value or upper.endswith(" C"):
                return round(numeric + 273.15, 2)
            if "K" in upper:
                return numeric
            if original_key and (
                "c" in original_key.lower() or "℃" in original_key or "摄氏" in original_key
            ) and "k" not in upper:
                return round(numeric + 273.15, 2)
        if isinstance(numeric, float):
            return int(numeric) if numeric.is_integer() else numeric
        return numeric
    if key in {"_energy_start_eV", "_energy_end_eV"}:
        numeric = _parse_numeric(value)
        if numeric is None:
            return value
        upper = value.upper()
        if "KEV" in upper:
            numeric = float(numeric) * 1_000
        if isinstance(numeric, float) and numeric.is_integer():
            return int(numeric)
        return numeric
    if (section, key) == ("sample", "mgid_list"):
        tokens = [
            token.strip()
            for token in re.split(r"[\s,;；，/\\]+", value)
            if token.strip()
        ]
        return tokens
    if (section, key) == ("files", "listing") and isinstance(value, str):
        return value.replace("\r\n", "\n").strip()
    return value


def _merge_into(
    metadata: MutableMapping[str, Dict[str, object]], section: str, key: str, value: object
) -> None:
    bucket = metadata.setdefault(section, {})
    existing = bucket.get(key)
    if isinstance(existing, list):
        new_items = value if isinstance(value, list) else [value]
        for item in new_items:
            if isinstance(item, str):
                candidate = item.strip()
            else:
                candidate = item
            if not candidate:
                continue
            if candidate not in existing:
                existing.append(candidate)
        return
    if isinstance(value, list):
        cleaned: List[object] = []
        for item in value:
            if isinstance(item, str):
                candidate = item.strip()
            else:
                candidate = item
            if candidate and candidate not in cleaned:
                cleaned.append(candidate)
        if cleaned:
            bucket[key] = cleaned
        return
    if isinstance(existing, str) and existing.strip():
        return
    if existing not in (None, ""):
        return
    bucket[key] = value


def _integrate_structured(metadata: MutableMapping[str, Dict[str, object]], data: Dict[str, object]) -> None:
    for section, payload in data.items():
        if section in metadata and isinstance(payload, dict):
            bucket = metadata.setdefault(section, {})
            for key, value in payload.items():
                if value in (None, ""):
                    continue
                if (section, key) == ("sample", "mgid_list") and isinstance(value, (list, tuple)):
                    cleaned = [
                        str(item).strip()
                        for item in value
                        if str(item).strip()
                    ]
                    if cleaned:
                        bucket[key] = cleaned
                    continue
                text = value if isinstance(value, str) else str(value)
                coerced = _coerce_value(section, key, text, original_key=key)
                bucket[key] = coerced
            continue

        direct_alias = TEMPLATE_FIELD_ALIASES.get(section)
        if direct_alias:
            target_section, target_key = direct_alias
            coerced = _coerce_value(target_section, target_key, str(payload), original_key=section)
            _merge_into(metadata, target_section, target_key, coerced)
            continue

        if isinstance(payload, dict):
            for sub_key, sub_value in payload.items():
                if sub_value in (None, ""):
                    continue
                alias = TEMPLATE_SECTION_FIELD_ALIASES.get((section, sub_key))
                if not alias:
                    continue
                target_section, target_key = alias
                coerced = _coerce_value(
                    target_section, target_key, str(sub_value), original_key=sub_key
                )
                _merge_into(metadata, target_section, target_key, coerced)


def _handle_cbf_special(raw_key: str, value: str) -> List[Tuple[str, str, str]]:
    normalized = _normalise_key(raw_key)
    if normalized in {"energy range", "energy range ev"}:
        numbers = re.findall(r"-?\d+(?:\.\d+)?(?:[eE][+-]?\d+)?", value)
        if len(numbers) >= 2:
            start = float(numbers[0])
            end = float(numbers[1])
            upper = value.upper()
            if "KEV" in upper:
                start *= 1_000.0
                end *= 1_000.0
            ordered = sorted((start, end))
            start_str = f"{ordered[0]:g}"
            end_str = f"{ordered[1]:g}"
            range_str = f"{ordered[0]:g}-{ordered[1]:g}"
            return [
                ("scan", "_energy_start_eV", start_str),
                ("scan", "_energy_end_eV", end_str),
                ("scan", "energy_range_eV", range_str),
            ]
    return []


def _integrate_cbf(metadata: MutableMapping[str, Dict[str, object]], path: Path) -> None:
    header_values = _parse_cbf_header(path)
    for raw_key, raw_value in header_values.items():
        value = raw_value.strip()
        if not value:
            continue
        special_entries = _handle_cbf_special(raw_key, value)
        if special_entries:
            for section, field, special_value in special_entries:
                coerced = _coerce_value(section, field, special_value, original_key=raw_key)
                _merge_into(metadata, section, field, coerced)
            continue
        transform = CBF_TRANSFORMS.get(raw_key.lower())
        if transform:
            transformed = transform(raw_value)
            if not transformed:
                continue
            value = transformed
        alias_key = _normalise_key(raw_key)
        if alias_key not in ALIASES:
            continue
        section, field = ALIASES[alias_key]
        coerced = _coerce_value(section, field, value, original_key=raw_key)
        _merge_into(metadata, section, field, coerced)


def _integrate_textual(metadata: MutableMapping[str, Dict[str, object]], path: Path) -> None:
    for key, value in _iter_text_key_values(path):
        alias_key = _normalise_key(key)
        if alias_key not in ALIASES:
            continue
        section, field = ALIASES[alias_key]
        coerced = _coerce_value(section, field, value, original_key=key)
        _merge_into(metadata, section, field, coerced)


def _integrate_tabular(metadata: MutableMapping[str, Dict[str, object]], path: Path) -> None:
    suffix = path.suffix.lower()
    if suffix in TABULAR_SUFFIXES_OPENPYXL:
        if load_workbook is None:  # pragma: no cover - optional dependency path
            raise RuntimeError(
                f"Cannot parse {path.name}: install openpyxl to read Excel metadata sheets"
            )
        try:
            workbook = load_workbook(path, data_only=True, read_only=True)
        except Exception as exc:  # pragma: no cover - delegated to caller
            raise RuntimeError(f"Failed to load Excel metadata {path.name}: {exc}") from exc
        try:
            sheet = workbook.active
            for row in sheet.iter_rows(values_only=True):
                if not row:
                    continue
                key_cell = row[0]
                if key_cell is None:
                    continue
                key = str(key_cell).strip()
                if not key:
                    continue
                values = [str(cell).strip() for cell in row[1:] if cell not in (None, "")]
                if not values:
                    continue
                value = " ".join(v for v in values if v)
                if not value:
                    continue
                alias_key = _normalise_key(key)
                if alias_key not in ALIASES:
                    continue
                section, field = ALIASES[alias_key]
                coerced = _coerce_value(section, field, value, original_key=key)
                _merge_into(metadata, section, field, coerced)
        finally:
            workbook.close()
        return

    if suffix in TABULAR_SUFFIXES_XLRD:
        if xlrd is None:  # pragma: no cover - optional dependency path
            raise RuntimeError(
                f"Cannot parse {path.name}: install xlrd to read legacy Excel metadata sheets"
            )
        try:
            workbook = xlrd.open_workbook(path, on_demand=True)
        except Exception as exc:  # pragma: no cover - delegated to caller
            raise RuntimeError(f"Failed to load Excel metadata {path.name}: {exc}") from exc
        try:
            sheet = workbook.sheet_by_index(0)
            for row_idx in range(sheet.nrows):
                key_cell = sheet.cell_value(row_idx, 0)
                if key_cell in (None, ""):
                    continue
                key = str(key_cell).strip()
                if not key:
                    continue
                values: List[str] = []
                for col_idx in range(1, sheet.ncols):
                    cell_value = sheet.cell_value(row_idx, col_idx)
                    if cell_value in (None, ""):
                        continue
                    values.append(str(cell_value).strip())
                if not values:
                    continue
                value = " ".join(v for v in values if v)
                if not value:
                    continue
                alias_key = _normalise_key(key)
                if alias_key not in ALIASES:
                    continue
                section, field = ALIASES[alias_key]
                coerced = _coerce_value(section, field, value, original_key=key)
                _merge_into(metadata, section, field, coerced)
        finally:
            workbook.release_resources()
        return

    raise RuntimeError(f"Unsupported Excel format: {path.suffix}")


def _integrate_logbook(metadata: MutableMapping[str, Dict[str, object]], path: Path) -> None:
    times: List[_dt.datetime] = []
    entries: List[str] = []
    for line in path.read_text(encoding="utf-8", errors="ignore").splitlines():
        line = line.strip()
        if not line:
            continue
        parts = line.split(maxsplit=2)
        if len(parts) >= 2:
            candidate = " ".join(parts[:2])
            dt = _parse_datetime(candidate)
            if dt:
                times.append(dt)
                if len(parts) == 3:
                    entries.append(parts[2])
                continue
        entries.append(line)
    if times:
        start = min(times)
        end = max(times)
        _merge_into(metadata, "experiment", "start_time", start.isoformat())
        _merge_into(metadata, "experiment", "end_time", end.isoformat())
    if entries:
        note = "; ".join(dict.fromkeys(entries))
        if note:
            bucket = metadata.setdefault("experiment", {})
            if not bucket.get("notes"):
                bucket["notes"] = note


def _extract_scan_statistics(path: Path) -> Tuple[int | None, Tuple[float, float] | None]:
    energies: List[float] = []
    try:
        if path.stat().st_size > 20_000_000:  # skip very large files
            return None, None
        text = path.read_text(encoding="utf-8", errors="ignore")
    except Exception:
        return None, None
    for raw_line in text.splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or line.startswith("//"):
            continue
        tokens = re.split(r"[\s,\t]+", line)
        if not tokens:
            continue
        try:
            energy = float(tokens[0])
        except ValueError:
            continue
        energies.append(energy)
    if not energies:
        return None, None
    energies.sort()
    points = len(energies)
    return points, (energies[0], energies[-1])


def _format_energy_range(bounds: Tuple[float, float]) -> str:
    start, end = bounds
    return f"{start:g}-{end:g}"


def _match_dataset_path(dataset_dir: Path, dataset_files: Iterable[Path], candidate: str) -> str:
    cleaned = candidate.strip().strip("\"\'")
    if not cleaned:
        return cleaned
    normalized = cleaned.replace("\\", "/")
    lowered = normalized.lower()
    files = dataset_files if isinstance(dataset_files, list) else list(dataset_files)
    for path in files:
        try:
            rel = path.relative_to(dataset_dir)
            rel_str = str(rel).replace("\\", "/")
        except ValueError:
            rel_str = None
        comparisons = {path.name.lower(), str(path.resolve()).lower()}
        if rel_str:
            comparisons.add(rel_str.lower())
        if lowered in comparisons:
            return rel_str if rel_str else path.name
    try:
        absolute_candidate = Path(cleaned)
    except Exception:
        absolute_candidate = None
    if absolute_candidate and absolute_candidate.is_absolute():
        resolved = absolute_candidate.resolve()
        for path in files:
            if path.resolve() == resolved:
                try:
                    return str(path.relative_to(dataset_dir))
                except ValueError:
                    return path.name
    if not Path(cleaned).is_absolute():
        resolved = (dataset_dir / cleaned).resolve()
        for path in files:
            if path.resolve() == resolved:
                try:
                    return str(path.relative_to(dataset_dir))
                except ValueError:
                    return path.name
    return normalized


def _normalise_listing_entries(
    entries: Iterable[str], dataset_dir: Path, dataset_files: Iterable[Path]
) -> List[str]:
    normalised: List[str] = []
    for raw in entries:
        candidate = raw.strip()
        if not candidate:
            continue
        resolved = _match_dataset_path(dataset_dir, dataset_files, candidate)
        if resolved and resolved not in normalised:
            normalised.append(resolved)
    return normalised


def _split_listing(value: str) -> List[str]:
    return [segment for segment in re.split(r"[\r\n;,；，]+", value) if segment.strip()]


def _gather_metadata(dataset_dir: Path, dataset_files: Iterable[Path]) -> Dict[str, Dict[str, object]]:
    file_list = list(dataset_files)
    metadata: Dict[str, Dict[str, object]] = {
        "experiment": {},
        "beamline": {},
        "detector": {},
        "sample": {},
        "scan": {},
        "files": {},
    }

    for file_path in file_list:
        suffix = file_path.suffix.lower()
        if suffix in STRUCTURED_SUFFIXES:
            try:
                data = _read_structured_metadata(file_path)
            except Exception as exc:
                print(
                    f"[synchrotron_metadata] skipped {file_path.name}: {exc}",
                    file=sys.stderr,
                )
                continue
            _integrate_structured(metadata, data)

    for file_path in file_list:
        suffix = file_path.suffix.lower()
        if suffix in CBF_SUFFIXES:
            try:
                _integrate_cbf(metadata, file_path)
            except Exception as exc:
                print(
                    f"[synchrotron_metadata] skipped {file_path.name}: {exc}",
                    file=sys.stderr,
                )
        elif suffix in TEXTUAL_SUFFIXES:
            if file_path.name.lower().startswith("logbook"):
                _integrate_logbook(metadata, file_path)
            else:
                _integrate_textual(metadata, file_path)
        elif suffix in TABULAR_SUFFIXES:
            try:
                _integrate_tabular(metadata, file_path)
            except Exception as exc:
                print(
                    f"[synchrotron_metadata] skipped {file_path.name}: {exc}",
                    file=sys.stderr,
                )

    files_bucket = metadata.setdefault("files", {})
    primary_value = files_bucket.get("primary")
    if isinstance(primary_value, str) and primary_value.strip():
        files_bucket["primary"] = _match_dataset_path(dataset_dir, file_list, primary_value)
    listing_value = files_bucket.get("listing")
    if isinstance(listing_value, str) and listing_value.strip():
        entries = _split_listing(listing_value)
        normalised_listing = _normalise_listing_entries(entries, dataset_dir, file_list)
        if normalised_listing:
            files_bucket["listing"] = "\n".join(normalised_listing)
        else:
            files_bucket["listing"] = listing_value.strip()
    elif isinstance(listing_value, list):
        entries = [str(item) for item in listing_value if str(item).strip()]
        if entries:
            normalised_listing = _normalise_listing_entries(entries, dataset_dir, file_list)
            if normalised_listing:
                files_bucket["listing"] = "\n".join(normalised_listing)

    # Provide sensible fallbacks if experiment/sample names were missing.
    default_name = dataset_dir.name
    if not metadata["experiment"].get("name"):
        metadata["experiment"]["name"] = default_name
    if not metadata["sample"].get("name"):
        metadata["sample"]["name"] = default_name

    scan_meta = metadata.get("scan", {})
    start_bound = scan_meta.pop("_energy_start_eV", None)
    end_bound = scan_meta.pop("_energy_end_eV", None)
    if not scan_meta.get("energy_range_eV"):
        if isinstance(start_bound, (int, float)) and isinstance(end_bound, (int, float)):
            ordered = sorted([float(start_bound), float(end_bound)])
            scan_meta["energy_range_eV"] = _format_energy_range((ordered[0], ordered[1]))

    # Derive the experiment date from the start time if available.
    start_time = metadata["experiment"].get("start_time")
    if isinstance(start_time, str):
        dt = _parse_datetime(start_time)
        if dt:
            metadata["experiment"]["start_time"] = dt.isoformat()

    needs_points = metadata["scan"].get("points") in (None, "", 0)
    needs_range = not metadata["scan"].get("energy_range_eV")
    if needs_points or needs_range:
        for candidate in file_list:
            suffix = candidate.suffix.lower()
            if suffix not in {".dat", ".txt", ".csv"}:
                continue
            if candidate.name.lower().startswith("logbook"):
                continue
            points, bounds = _extract_scan_statistics(candidate)
            if needs_points and points:
                metadata["scan"]["points"] = points
                needs_points = False
            if needs_range and bounds:
                metadata["scan"]["energy_range_eV"] = _format_energy_range(bounds)
                needs_range = False
            if not needs_points and not needs_range:
                break

    primary = metadata.get("files", {}).get("primary")
    if not primary:
        candidate = _choose_primary_file(dataset_dir, file_list)
        if candidate:
            metadata.setdefault("files", {})["primary"] = candidate

    return metadata


def _choose_primary_file(dataset_dir: Path, dataset_files: Iterable[Path]) -> str | None:
    files = list(dataset_files)
    for suffix in DATA_PRIORITIES:
        for path in files:
            if path.suffix.lower() == suffix:
                try:
                    return str(path.relative_to(dataset_dir))
                except ValueError:
                    return path.name
    if files:
        try:
            return str(files[0].relative_to(dataset_dir))
        except ValueError:
            return files[0].name
    return None


def populate_template(
    template: Dict[str, object],
    metadata: Dict[str, object],
    dataset_dir: Path,
    dataset_files: Iterable[Path],
) -> Dict[str, object]:
    filled = deepcopy(template)

    experiment = metadata.get("experiment", {}) if isinstance(metadata.get("experiment"), dict) else {}
    beamline = metadata.get("beamline", {}) if isinstance(metadata.get("beamline"), dict) else {}
    detector = metadata.get("detector", {}) if isinstance(metadata.get("detector"), dict) else {}
    sample = metadata.get("sample", {}) if isinstance(metadata.get("sample"), dict) else {}
    scan = metadata.get("scan", {}) if isinstance(metadata.get("scan"), dict) else {}
    files = metadata.get("files", {}) if isinstance(metadata.get("files"), dict) else {}

    _update(filled, "实验名称", experiment.get("name", ""))
    _update(filled, "实验负责人", experiment.get("principal_investigator", ""))
    _update(filled, "实验团队", experiment.get("team", ""))
    _update(filled, "实验地点", experiment.get("location", ""))

    mgid_custom = experiment.get("mgid_custom")
    if isinstance(mgid_custom, str):
        filled["MGID自定义部分"] = mgid_custom.strip()
    elif mgid_custom not in (None, ""):
        filled["MGID自定义部分"] = str(mgid_custom)
    else:
        filled["MGID自定义部分"] = str(filled.get("MGID自定义部分", "")).strip()

    start_time = _parse_datetime(str(experiment.get("start_time", ""))) if experiment else None
    if start_time is not None:
        _update(filled, "实验日期", start_time.date().isoformat())
    notes = experiment.get("notes", "")
    if notes:
        existing_note = str(filled.get("实验备注", "")).strip()
        filled["实验备注"] = notes if not existing_note else f"{existing_note}; {notes}"

    beam_section = dict(filled.get("光束参数", {}))
    _update(beam_section, "光源设施名称", beamline.get("facility", ""))
    _update(beam_section, "束线名称", beamline.get("beamline", ""))
    if beamline.get("photon_energy_eV") is not None:
        beam_section["光子能量(eV)"] = beamline.get("photon_energy_eV")
    if beamline.get("storage_ring_current_mA") is not None:
        beam_section["储存环电流(mA)"] = beamline.get("storage_ring_current_mA")
    _update(beam_section, "单色器", beamline.get("monochromator", ""))
    slit_width = beamline.get("exit_slit_um")
    if slit_width not in (None, ""):
        beam_section["出射狭缝宽度(μm)"] = slit_width
    filled["光束参数"] = beam_section

    detector_section = dict(filled.get("探测系统", {}))
    _update(detector_section, "探测器名称", detector.get("name", ""))
    _update(detector_section, "探测器型号", detector.get("model", ""))
    if detector.get("distance_mm") is not None:
        detector_section["样品到探测器距离(mm)"] = detector.get("distance_mm")
    filled["探测系统"] = detector_section

    sample_section = dict(filled.get("样品信息", {}))
    _update(sample_section, "样品名称", sample.get("name", ""))
    _update(sample_section, "样品环境", sample.get("environment", ""))
    if sample.get("temperature_K") is not None:
        sample_section["样品温度(K)"] = sample.get("temperature_K")
    filled["样品信息"] = sample_section

    sample_mgids = sample.get("mgid_list")
    mgid_values: List[str] = []
    if isinstance(sample_mgids, str):
        sample_mgids = [sample_mgids]
    if isinstance(sample_mgids, list):
        for item in sample_mgids:
            text = str(item).strip()
            if text and text not in mgid_values:
                mgid_values.append(text)
    filled["关联样品MGID"] = mgid_values

    scan_section = dict(filled.get("扫描参数", {}))
    _update(scan_section, "扫描模式", scan.get("mode", ""))
    if scan.get("points") is not None:
        scan_section["扫描点数"] = scan.get("points")
    if scan.get("dwell_time_s") is not None:
        scan_section["单点积分时间(s)"] = scan.get("dwell_time_s")
    _update(scan_section, "扫描能区范围(eV)", scan.get("energy_range_eV", ""))
    filled["扫描参数"] = scan_section

    comments = scan.get("comments", "")
    if comments:
        existing = str(filled.get("备注", "")).strip()
        filled["备注"] = comments if not existing else f"{existing}; {comments}"

    paths = sorted(Path(p) for p in dataset_files)
    original_files = dict(filled.get("原始文件", {}))
    if paths:
        primary = files.get("primary") if isinstance(files, dict) else None
        if not primary:
            primary = _choose_primary_file(dataset_dir, paths)
        if primary:
            original_files["表征原始数据文件"] = primary
        listing_value = files.get("listing") if isinstance(files, dict) else None
        if listing_value:
            original_files["表征原始数据列表"] = str(listing_value).strip()
        else:
            rel_lines: List[str] = []
            for p in paths:
                try:
                    rel = p.relative_to(dataset_dir)
                except ValueError:
                    rel = p
                rel_lines.append(str(rel))
            original_files["表征原始数据列表"] = "\n".join(rel_lines)
    filled["原始文件"] = original_files

    return filled


def generate_metadata(
    input_path: Path | str,
    *,
    template_path: Path | str | None = None,
    output_path: Path | str | None = None,
) -> Dict[str, object]:
    """Return a populated synchrotron radiation template for ``input_path``."""

    target = Path(input_path).expanduser().resolve()
    dataset_dir, dataset_files = _collect_dataset_files(target)
    dataset_list = [Path(p) for p in dataset_files]

    if output_path is not None:
        output_resolved = Path(output_path).expanduser().resolve()
        dataset_list = [p for p in dataset_list if p.resolve() != output_resolved]

    metadata = _gather_metadata(dataset_dir, dataset_list)
    template_file = (
        Path(template_path).expanduser().resolve()
        if template_path is not None
        else DEFAULT_TEMPLATE
    )
    template_data = json.loads(template_file.read_text("utf-8"))
    return populate_template(template_data, metadata, dataset_dir, dataset_list)


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "input",
        type=Path,
        help="Directory that contains the synchrotron data or the path to a metadata file",
    )
    parser.add_argument(
        "--template",
        type=Path,
        default=DEFAULT_TEMPLATE,
        help="Path to the synchrotron radiation metadata template (JSON)",
    )
    parser.add_argument(
        "--output",
        type=Path,
        help="Write the populated template to this file instead of stdout",
    )
    parser.add_argument(
        "--pretty",
        action="store_true",
        help="Pretty-print JSON with indentation",
    )
    args = parser.parse_args()

    filled = generate_metadata(
        args.input,
        template_path=args.template,
        output_path=args.output,
    )
    json_text = json.dumps(filled, ensure_ascii=False, indent=2 if args.pretty else None)

    if args.output:
        args.output.write_text(json_text + ("\n" if args.pretty else ""), encoding="utf-8")
    else:
        print(json_text)


if __name__ == "__main__":
    main()
